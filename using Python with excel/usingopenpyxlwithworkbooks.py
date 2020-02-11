from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = Workbook()
ws = wb.active
# create a new workbook with 2 worksheet
ws1 = wb.create_sheet('NewSheet')
ws2 = wb.create_sheet('NewSheet1',0)
ws.title='MySheet'

print(wb.sheetnames)

wb2 = load_workbook('modified.xlsx')
new_sheet = wb2.create_sheet('NewSheet')
active_sheet = wb2.active

cell = active_sheet['A1']
print(cell)
active_sheet['A1'] = 0
wb2.save('modified1.xlsx')
print(cell.value)



#df = pd.read_csv('Profile.csv',header=None)
#df.columns = ['Items','Brand','price','earn','a','b','c']