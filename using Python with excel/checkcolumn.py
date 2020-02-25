from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd	
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import xlrd
import numpy as np
import matplotlib.pyplot as plt

#=============================================================
#testNumber = int(input('How many tests have run : '))
testNumber = 25
#dataNumber = int(input('How many data per test : '))
dataNumber = 5
#fid = input('File name : (with extension)')
fid = 'consistancy test results.xlsx'
wb = load_workbook(fid)
ws = wb.active
print(ws.max_row)
print(ws.max_column)