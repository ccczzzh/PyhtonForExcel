from openpyxl.workbook import Workbook
from openpyxl import load_workbook
import pandas as pd	
import xlsxwriter
from xlsxwriter.utility import xl_rowcol_to_cell
import xlrd
import numpy as np
import matplotlib.pyplot as plt

#=============================================================
#testNumber = int(input('How many tests have run : '))
testNumber = 25
#dataNumber = int(input('How many data per test : '))
dataNumber = 5
#fid = input('File name : (with extension)')
fid = 'consistancy test results.xlsx'
wb = load_workbook(fid)
ws = wb.active
row_count = ws.max_row
col_count = ws.max_column
mytable = np.zeros((testNumber*dataNumber,1))
a = 0
for ws in xlrd.open_workbook(fid).sheets():
	for row in range(ws.nrows):
		for col in range(ws.ncols):
			myCell = ws.cell(row,col)
			#print(len(row))
			if myCell.value == 'Act':
				cvn = 1
				while 
				if isinstance(ws.cell((row+cvn),col).value, None)
					cvn =+ 1
					b_c = cvn + 1
					mytable[a] = ws.cell((row+cvn),col).value
					#nextvalue = ws.cell((row+b_c),col).value
					while row_c < row_count:
						if isinstance(ws.cell((row+b_c),col).value,float):
							a += 1
							mytable[a] = ws.cell((row+b_c),col).value
							#print(b_c)
						else:
							break
						b_c += 1
				a += 1
#print(mytable)
newtable=mytable.reshape((testNumber,dataNumber))

#print(newtable)

df = pd.DataFrame(newtable)
df.to_excel(excel_writer = 'ReadyData.xlsx')
# newtable1 = newtable.transpose()
# print(newtable1)
# aver = np.average(newtable1,axis=1)
# print(aver[0])
# stde = np.std(newtable1,axis=1)
# print(stde)
# tol = 0.0001
# tol = input('tolerance = ')
# for i in range(0,dataNumber):
	# horiz_line_data = np.array([aver[i] for j in range(testNumber)])
	# plt.plot(newtable1[i],'-x',horiz_line_data*(1+tol),'b-o',horiz_line_data*(1-tol),'g-o')
	# plt.show()
	# #figid = 'g.png'
	# figid = input('Enter the figure Name: (with extension name) ')
	# plt.savefig(figid)