from openpyxl.styles import Font,colors,Color,Alignment,PatternFill,GradientFill,Border,Side
from openpyxl.styles import NamedStyle

from openpyxl.workbook import Workbook

wb=Workbook()
ws = wb.active

for i in range(1,20):
    ws.append(range(300))

ws.merge_cells('A1:B5')
ws.unmerge_cells("A1:B5")
ws.merge_cells(start_row = 2, start_column =2,end_row =5, end_column=5)

cell=ws["B2"]

cell.font = Font(color=colors.RED,size=20,italic=True)
cell.value = 'Merged Cell'
cell.alignment =Alignment(horizontal='right',vertical='bottom')
cell.fill=GradientFill(stop=("000000","FFFFFF"))

wb.save('text.xlsx')

highlight = NamedStyle(name='highlight')
highlight.font = Font(bold=True)
bd = Side(style='thick',color='000000')
highlight.border = Border(left=bd,top=bd,right=bd,bottom=bd)
highlight.fill=PatternFill('solid',fgColor='FFFF00')

count =0
for col in ws.iter_cols(min_col=8,min_row=1,max_col=30,max_row=30):
    col[count].style = highlight
    count +=1

wb.save('highlight.xlsx')