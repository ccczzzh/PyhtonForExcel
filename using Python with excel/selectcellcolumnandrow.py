from openpyxl.workbook import Workbook
from openpyxl import load_workbook

wb = load_workbook('modified.xlsx')
ws = wb.active
#print cell range
cell_range=ws['A1':'C1']
#print(cell_range)#print the cell range not in matrix format
# print column range
col_c = ws['C']
col_range = ws['A':'C']
#print(col_range)

#pring row range

row_range = ws[1:4]
#print(row_range)

for row in ws.iter_rows(min_row=1, max_col=2, max_row=4, values_only=True):
    for cell in row:
        print(cell)
