def search():
	from openpyxl.workbook import Workbook
	from openpyxl import load_workbook
	import pandas as pd	
	import xlsxwriter
	from xlsxwriter.utility import xl_rowcol_to_cell
	import xlrd
	import numpy as np
	wb = load_workbook('consistancy test results.xlsx')
	ws = wb.active
	mytable = np.zeros((100,1))
	a = 0
	for ws in xlrd.open_workbook('consistancy test results.xlsx').sheets():
		for row in range(ws.nrows):
			for col in range(ws.ncols):
				myCell = ws.cell(row,col)

				if myCell.value == 'Act':
					
					# print('========')
					# print(xl_rowcol_to_cell(row,col))
					mytable[a] = ws.cell((row+1),col).value
					#print(ws.cell((row+1),col).value)
					a +=1